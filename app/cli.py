import argparse
import os
import sys
import time
from app.orchestrator import run_all, run_yahoo, run_nst, run_merge
from diagnostics.runner import run_dq as run_dq_diag, run_dq_prior as run_dq_prior_diag
from application.forecast import forecast as run_forecast
from application.compare import compare as run_compare
from application.analyze import evaluate as run_analyze
from application.waiver_agent import recommend as run_waiver

# If you store canonical inputs in known NSTstats paths, set sensible defaults here
DEFAULT_OUT_CSV = "data/lookup_table.csv"


def cmd_schedule_refresh(argv=None):
    p = argparse.ArgumentParser(
        prog="nststats schedule-lookup",
        description="Build NHL schedule lookup using NHLschedule defaults; output saved to data/lookup_table.csv"
    )
    p.add_argument("--out-csv", default=DEFAULT_OUT_CSV,
                   help="Output CSV path (default: data/lookup_table.csv)")
    p.add_argument("--refresh-cache", action="store_true",
                   help="Bypass cache and refetch all team tables (proxied to nhl_schedule)")
    p.add_argument(
        "--nhl-schedule-path",
        dest="nhl_schedule_path",
        default=None,
        help=(
            "Optional path to your local NHLschedule project. If provided (or if the NHL_SCHEDULE_PATH env var is set),\n"
            "we will import nhl_schedule from there so recent local edits are used without reinstalling."
        ),
    )

    args = p.parse_args(argv)

    # Ensure output directory exists
    out_csv = args.out_csv or DEFAULT_OUT_CSV
    out_dir = os.path.dirname(out_csv)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    # If user passed a local path (or env var is set), insert it into sys.path for this invocation
    nhl_schedule_path = args.nhl_schedule_path or os.environ.get("NHL_SCHEDULE_PATH")
    if nhl_schedule_path and os.path.isdir(nhl_schedule_path):
        if nhl_schedule_path not in sys.path:
            sys.path.insert(0, nhl_schedule_path)

    # Import here so that sys.path injection (if any) takes effect
    try:
        from nhl_schedule.build_lookup import build
    except Exception as e:
        raise RuntimeError(
            "Failed to import nhl_schedule. If you are developing NHLschedule in a separate project, "
            "either install it (pip install -e <path>) or pass --nhl-schedule-path / set NHL_SCHEDULE_PATH."
        ) from e

    # Use NHLschedule defaults for schedule path and sheet/table by passing None
    out_path = build(
        schedule_path=None,
        sheet_or_table=None,
        out_csv=out_csv,
        out_xlsx=None,
        refresh_cache=args.refresh_cache,
    )
    print(f"Lookup table written to: {out_path}")

def main():
    parser = argparse.ArgumentParser(description="NSTstats CLI")
    # Top-level flag alias so users can run: python -m app.cli --schedule-lookup
    parser.add_argument("--schedule-lookup", dest="schedule_lookup", action="store_true", help="Run schedule lookup (alias for the 'schedule-lookup' subcommand)")

    subparsers = parser.add_subparsers(dest="command")

    subparsers.add_parser("all", help="Run Yahoo and NST pipelines")
    subparsers.add_parser("yahoo", help="Run Yahoo Fantasy helper")
    nst = subparsers.add_parser("nst", help="Run NST pipelines (skaters, goalies)")
    nst.add_argument("--refresh-prior", dest="refresh_prior", action="store_true", help="Force re-fetch/re-score prior-season data even if cached CSVs exist")
    nst.add_argument("--skip-prior", dest="skip_prior", action="store_true", help="Skip prior-season fetch/scoring regardless of cache state")
    subparsers.add_parser("merge", help="Merge NST with Yahoo ownership to CSVs")

    # Average comparison grid (team vs opp and league) from DB
    avg_cmp = subparsers.add_parser(
        "avg-compare",
        help="Print average stat values prior to current week: Team vs Opponent and Team vs League Avg (split by position type)",
    )
    avg_cmp.add_argument("--league-key", dest="league_key", required=True, help="Yahoo league_key (e.g., nhl.p.2526)")
    avg_cmp.add_argument("--current-week", dest="current_week", type=int, required=True, help="Current week number; averages use weeks < current_week")
    avg_cmp.add_argument("--team-id", dest="team_id", required=True, help="Team identifier (numeric id like 4 or full team_key)")
    avg_cmp.add_argument("--opp-team-id", dest="opp_team_id", required=True, help="Opponent team identifier (numeric id like 8 or full team_key)")

    # Fetch per-date roster and player GP into RosterSlotDaily (Yahoo API; rate-limit friendly)
    fd = subparsers.add_parser(
        "fetch-daily-gp",
        help="Populate RosterSlotDaily by fetching team rosters and per-date player stats for a week range",
    )
    fd.add_argument("--league-key", dest="league_key", required=True, help="Yahoo league_key (e.g., nhl.l.12345)")
    fd.add_argument("--start-week", dest="start_week", type=int, required=True, help="Start week number (inclusive)")
    fd.add_argument("--end-week", dest="end_week", type=int, required=True, help="End week number (inclusive)")
    fd.add_argument(
        "--team-keys",
        dest="team_keys",
        default=None,
        help="Optional comma-separated list of team_keys to restrict processing (e.g., nhl.l.XXXXX.t.1,nhl.l.XXXXX.t.4)",
    )
    fd.add_argument("--sleep-sec", dest="sleep_sec", type=float, default=1.0, help="Seconds to sleep between team calls (default 1.0)")
    fd.add_argument("--dry-run", dest="dry_run", action="store_true", help="Do not write to DB; print planned actions only")

    # Backfill WeeklyPlayerGP from daily roster slots (DB-only; no Yahoo API calls)
    bf = subparsers.add_parser(
        "backfill-gp",
        help="Aggregate daily roster slots into WeeklyPlayerGP for a week range and optionally close weeks",
    )
    bf.add_argument("--league-key", dest="league_key", required=True, help="Yahoo league_key (e.g., nhl.l.12345)")
    bf.add_argument("--start-week", dest="start_week", type=int, default=1, help="Start week number (default 1)")
    bf.add_argument("--end-week", dest="end_week", type=int, default=8, help="End week number inclusive (default 8)")
    bf.add_argument(
        "--team-keys",
        dest="team_keys",
        default=None,
        help="Optional comma-separated list of team_keys to restrict processing (e.g., nhl.p.2526.t.1,nhl.p.2526.t.4)",
    )
    bf.add_argument("--batch-size", dest="batch_size", type=int, default=3, help="Teams per batch before sleeping (default 3)")
    bf.add_argument("--sleep-sec", dest="sleep_sec", type=float, default=2.0, help="Seconds to sleep between batches (default 2.0)")
    bf.add_argument("--force", dest="force", action="store_true", help="Overwrite existing WeeklyPlayerGP rows (default: skip existing)")
    bf.add_argument("--no-close", dest="no_close", action="store_true", help="Do not mark weeks closed after backfill")
    bf.add_argument("--dry-run", dest="dry_run", action="store_true", help="Do not write to DB; print planned actions only")

    # Schedule lookup command (nhl_schedule integration)
    sched = subparsers.add_parser("schedule-lookup", help="Build or update the weekly schedule lookup table from inputs")
    # sched.add_argument("--matchups", default=DEFAULT_MATCHUPS, help="Path to matchups parquet/csv with team vs opponent per date")
    # sched.add_argument("--opp-ease", dest="opp_ease", default=DEFAULT_OPP_EASE, help="Path to opponent ease parquet/csv with OppDefenseScore0to100 by team")
    # sched.add_argument("--out-csv", dest="out_csv", default=DEFAULT_OUT_CSV, help="Output CSV path for the lookup table")
    # sched.add_argument("--out-xlsx", dest="out_xlsx", default=DEFAULT_OUT_XLSX, help="Optional Excel output path")

    # Init weeks from nhl_schedule (populate Week table start/end dates)
    iw = subparsers.add_parser(
        "init-weeks",
        help="Upsert Week rows (start/end dates) for a league using nhl_schedule (Yahoo week windows)",
    )
    iw.add_argument("--league-key", dest="league_key", required=True, help="Yahoo league_key (e.g., nhl.p.2526)")
    iw.add_argument("--season", dest="season", required=False, help="Season year (e.g., 2025). Defaults to SEASON in .env")
    iw.add_argument("--start-week", dest="start_week", type=int, default=10, help="First week to load (default 10)")
    iw.add_argument("--end-week", dest="end_week", type=int, default=24, help="Last week to load (default 24)")
    iw.add_argument("--refresh-cache", dest="refresh_cache", action="store_true", help="Force refresh of nhl_schedule cached data")
    iw.add_argument("--source", dest="source", choices=["auto", "lookup", "excel"], default="auto", help="Where to source weeks: nhl_schedule lookup CSV, Excel sheet, or auto (try both)")
    iw.add_argument("--excel-path", dest="excel_path", default=None, help="Path to Excel schedule (e.g., AG_2526_Schedule.xlsx)")
    iw.add_argument("--excel-sheet", dest="excel_sheet", default="Yahoo_Weeks", help="Excel sheet name containing Yahoo weeks (default Yahoo_Weeks)")
    iw.add_argument("--debug", dest="debug", action="store_true", help="Print detected headers and sample parsed rows for troubleshooting")
    iw.add_argument("--dry-run", dest="dry_run", action="store_true", help="Do not write to DB; print planned upserts only")

    # Diagnostics command
    dq = subparsers.add_parser("dq", help="Run distribution diagnostics on skaters and goalies")
    dq.add_argument("--windows", nargs="*", default=["szn"], help="Windows to analyze (e.g., szn l7)")
    dq.add_argument("--metrics", nargs="*", default=["G","A","PPP","SOG","FOW","HIT","BLK","PIM"], help="Metrics to analyze")
    dq.add_argument("--min-n", type=int, default=25, help="Minimum sample size per group")
    dq.add_argument("--out", default=None, help="Output directory root (default data/dq)")
    dq.add_argument("--prior-csv", dest="prior_csv", default=None, help="Optional prior-season CSV to compare (e.g., data/merged_skaters_prior.csv)")

    # Forecast command
    fc = subparsers.add_parser("forecast", help="Compute per-player stat forecasts for ROW, Next Week, and ROS")
    fc.add_argument("--current-week", dest="current_week", type=int, required=True, help="Current fantasy week number (1..24)")
    fc.add_argument("--season-weight", dest="season_weight", type=float, default=0.5, help="Weight for season-to-date rates (manual override; ignored if auto-weights is enabled)")
    fc.add_argument("--last7-weight", dest="last7_weight", type=float, default=0.2, help="Weight for last-7 games rates (default 0.2; held constant when auto-weights is enabled)")
    fc.add_argument("--last-year-weight", dest="last_year_weight", type=float, default=0.3, help="Weight for last-year per-60 based per-game expectation (manual override; ignored if auto-weights is enabled)")
    fc.add_argument("--sos-weight", dest="sos_weight", type=float, default=0.3, help="Schedule strength impact weight; positive means hard schedule lowers output (default 0.3)")
    fc.add_argument("--horizons", nargs="*", default=["row","next","ros"], help="Horizons to compute: row next ros")
    fc.add_argument("--skaters-csv", dest="skaters_csv", default=os.path.join("data","merged_skaters.csv"), help="Input merged skaters CSV (default data/merged_skaters.csv)")
    fc.add_argument("--lookup-csv", dest="lookup_csv", default=os.path.join("data","lookup_table.csv"), help="Schedule lookup CSV (default data/lookup_table.csv)")
    fc.add_argument("--out-csv", dest="out_csv", default=os.path.join("data","forecasts.csv"), help="Output CSV path (default data/forecasts.csv)")
    fc.add_argument("--auto-weights", dest="auto_weights", action="store_true", default=True, help="Enable sliding-scale weights between last-year and season-to-date (default on)")
    fc.add_argument("--no-auto-weights", dest="auto_weights", action="store_false", help="Disable sliding-scale weights; use manual season/last7/last-year weights")
    fc.add_argument("--weeks-in-season", dest="weeks_in_season", type=int, default=24, help="Total number of fantasy weeks (default 24)")

    # Compare command
    cp = subparsers.add_parser("compare", help="Compare our forecasts vs NatePts projections and Last-Year benchmarks")
    cp.add_argument("--current-week", dest="current_week", type=int, required=True, help="Current fantasy week number (1..24)")
    cp.add_argument("--forecast-csv", dest="forecast_csv", default=os.path.join("data","forecasts.csv"), help="Forecast CSV input (default data/forecasts.csv)")
    cp.add_argument("--lookup-csv", dest="lookup_csv", default=os.path.join("data","lookup_table.csv"), help="Schedule lookup CSV (default data/lookup_table.csv)")
    cp.add_argument("--proj-xlsx", dest="proj_xlsx", default=os.path.join("NatePts.xlsx"), help="Path to NatePts.xlsx (default NatePts.xlsx in project root)")
    cp.add_argument("--proj-sheet", dest="proj_sheet", default="NatePts", help="Sheet/Table name in NatePts.xlsx (default NatePts)")
    cp.add_argument("--ly-sit-s", dest="ly_sit_s_csv", default=None, help="Optional CSV of NST Last-Year all-situations per-60 (rate=y)")
    cp.add_argument("--ly-sit-pp", dest="ly_sit_pp_csv", default=None, help="Optional CSV of NST Last-Year power-play per-60 (rate=y)")
    cp.add_argument("--horizons", nargs="*", default=["row","next","ros"], help="Horizons to compute: row next ros")
    cp.add_argument("--all-players", dest="all_players", action="store_true", help="Include all players (default: owned only)")
    cp.add_argument("--out-csv", dest="out_csv", default=os.path.join("data","compare.csv"), help="Output CSV path (default data/compare.csv)")

    # Analyze command
    az = subparsers.add_parser("analyze", help="Evaluate season-total forecasts vs projections and export metrics/Excel")
    az.add_argument("--compare-csv", dest="compare_csv", default=os.path.join("data","compare.csv"), help="Input compare CSV (default data/compare.csv)")
    az.add_argument("--out-dir", dest="out_dir", default=os.path.join("data","eval"), help="Output directory for metrics/Excel (default data/eval)")

    # Waiver agent command
    wa = subparsers.add_parser("waiver-agent", help="Recommend waiver-wire streamers vs your roster by position and horizon")
    wa.add_argument("--team-name", dest="team_name", required=True, help="Exact team_name to analyze (e.g., Trocheck Backcheck Paycheque)")
    wa.add_argument("--current-week", dest="current_week", type=int, required=True, help="Current fantasy week number (1..24)")
    wa.add_argument("--skaters-csv", dest="skaters_csv", default=os.path.join("data","merged_skaters.csv"), help="Merged skaters CSV (default data/merged_skaters.csv)")
    wa.add_argument("--lookup-csv", dest="lookup_csv", default=os.path.join("data","lookup_table.csv"), help="Schedule lookup CSV (default data/lookup_table.csv)")
    wa.add_argument("--prior-csv", dest="prior_csv", default=os.path.join("data","merged_skaters_prior.csv"), help="Prior-season merged skaters CSV (default data/merged_skaters_prior.csv)")
    wa.add_argument("--out-dir", dest="out_dir", default=os.path.join("output"), help="Output directory for CSV (default output/)")

    args = parser.parse_args()

    # Determine command, honoring the top-level flag alias if provided
    if getattr(args, "schedule_lookup", False) and not getattr(args, "command", None):
        cmd = "schedule-lookup"
    else:
        cmd = args.command or "all"

    if cmd == "yahoo":
        run_yahoo()
    elif cmd == "nst":
        run_nst(
            refresh_prior=getattr(args, "refresh_prior", False),
            skip_prior=getattr(args, "skip_prior", False),
        )
    elif cmd == "merge":
        run_merge()
    elif cmd == "dq":
        windows = getattr(args, "windows", ["szn"]) or ["szn"]
        metrics = getattr(args, "metrics", ["G","A","PPP","SOG","FOW","HIT","BLK","PIM"]) or ["G","A","PPP","SOG","FOW","HIT","BLK","PIM"]
        min_n = getattr(args, "min_n", 25)
        out = getattr(args, "out", None)
        out_dir = out if out else None
        prior_csv = getattr(args, "prior_csv", None)
        # 1) Skaters (F/D)
        out_path = run_dq_diag(
            input_csv=os.path.join("data", "skaters_scored.csv"),
            out_dir=out_dir or os.path.join("data", "dq"),
            metrics=metrics,
            windows=windows,
            min_n=min_n,
            prior_input_csv=None,  # keep current season separate from prior outputs
            group_by_position=True,
        )
        # 2) Goalies (G) — append to same summary
        try:
            g_metrics = ["GA", "SV%", "GAA"]
            g_csv = os.path.join("data", "merged_goalies.csv")
            if os.path.exists(g_csv):
                run_dq_diag(
                    input_csv=g_csv,
                    out_dir=out_dir or os.path.join("data", "dq"),
                    metrics=g_metrics,
                    windows=windows,
                    min_n=min_n,
                    prior_input_csv=None,
                    group_by_position=False,
                    segments=["G"],
                    existing_root=out_path,
                )
        except Exception as _e:
            print(f"[Warn] Goalie diagnostics skipped due to error: {_e}")
        # 3) Prior season runs in a separate dq/prior/<ts> directory
        try:
            ts = os.path.basename(out_path.rstrip(os.sep))
            # Prior Skaters
            prior_skaters_path = prior_csv or os.path.join("data", "merged_skaters_prior.csv")
            if os.path.exists(prior_skaters_path):
                run_dq_prior_diag(
                    input_csv=prior_skaters_path,
                    out_dir=os.path.join("data", "dq", "prior"),
                    metrics=metrics,
                    windows=None,  # default to ["szn"] for prior
                    min_n=min_n,
                    group_by_position=True,
                    segments=None,
                    existing_ts=ts,
                )
            # Prior Goalies
            prior_goalies_path = os.path.join("data", "merged_goalies_prior.csv")
            if os.path.exists(prior_goalies_path):
                run_dq_prior_diag(
                    input_csv=prior_goalies_path,
                    out_dir=os.path.join("data", "dq", "prior"),
                    metrics=g_metrics,
                    windows=None,  # default to ["szn"] for prior
                    min_n=min_n,
                    group_by_position=False,
                    segments=["G"],
                    existing_ts=ts,
                )
        except Exception as _e:
            print(f"[Warn] Prior-season diagnostics skipped due to error: {_e}")
        print(f"Diagnostics written to: {out_path}")
    elif cmd == "forecast":
        current_week = getattr(args, "current_week")
        season_weight = getattr(args, "season_weight", 0.8)
        last7_weight = getattr(args, "last7_weight", 0.2)
        sos_weight = getattr(args, "sos_weight", 0.3)
        horizons = tuple(getattr(args, "horizons", ["row","next","ros"]))
        skaters_csv = getattr(args, "skaters_csv", os.path.join("data","merged_skaters.csv"))
        lookup_csv = getattr(args, "lookup_csv", os.path.join("data","lookup_table.csv"))
        out_csv = getattr(args, "out_csv", os.path.join("data","forecasts.csv"))
        out_path = run_forecast(
            skaters_csv=skaters_csv,
            lookup_csv=lookup_csv,
            out_csv=out_csv,
            current_week=current_week,
            season_weight=season_weight,
            last7_weight=last7_weight,
            last_year_weight=getattr(args, "last_year_weight", 0.0),
            sos_weight=sos_weight,
            horizons=horizons,
            auto_weights=getattr(args, "auto_weights", True),
            weeks_in_season=getattr(args, "weeks_in_season", 24),
        )
        print(f"Forecasts written to: {out_path}")
    elif cmd == "schedule-lookup":
        # Pass an empty argv so the inner parser uses defaults and doesn't try to parse the outer Namespace
        cmd_schedule_refresh(argv=[])
    elif cmd == "compare":
        current_week = getattr(args, "current_week")
        forecast_csv = getattr(args, "forecast_csv", os.path.join("data","forecasts.csv"))
        lookup_csv = getattr(args, "lookup_csv", os.path.join("data","lookup_table.csv"))
        proj_xlsx = getattr(args, "proj_xlsx", os.path.join("NatePts.xlsx"))
        proj_sheet = getattr(args, "proj_sheet", "NatePts")
        ly_sit_s_csv = getattr(args, "ly_sit_s_csv", None)
        ly_sit_pp_csv = getattr(args, "ly_sit_pp_csv", None)
        horizons = tuple(getattr(args, "horizons", ["row","next","ros"]))
        out_csv = getattr(args, "out_csv", os.path.join("data","compare.csv"))
        all_players = bool(getattr(args, "all_players", False))
        out_path = run_compare(
            current_week=current_week,
            forecast_csv=forecast_csv,
            lookup_csv=lookup_csv,
            proj_xlsx=proj_xlsx,
            proj_sheet=proj_sheet,
            ly_sit_s_csv=ly_sit_s_csv,
            ly_sit_pp_csv=ly_sit_pp_csv,
            horizons=horizons,
            out_csv=out_csv,
            all_players=all_players,
        )
        print(f"Comparison written to: {out_path}")
    elif cmd == "analyze":
        compare_csv = getattr(args, "compare_csv", os.path.join("data", "compare.csv"))
        out_dir = getattr(args, "out_dir", os.path.join("data", "eval"))
        out_path = run_analyze(compare_csv=compare_csv, out_dir=out_dir)
        print(f"Season total evaluation written to: {out_path}\nPlots and CSVs in: {out_dir}")
    elif cmd == "avg-compare":
        from application.avg_compare import compare_averages as run_avg_compare
        league_key = getattr(args, "league_key")
        current_week = int(getattr(args, "current_week"))
        team_id = str(getattr(args, "team_id"))
        opp_team_id = str(getattr(args, "opp_team_id"))
        run_avg_compare(
            league_key=league_key,
            current_week=current_week,
            team_id=team_id,
            opp_team_id=opp_team_id,
        )
    elif cmd == "backfill-gp":
        # Lazy import SQLAlchemy models/helpers to keep CLI import-time light
        from sqlalchemy import func
        from infrastructure.persistence import (
            get_session,
            League,
            Week,
            Team,
            RosterSlotDaily,
            WeeklyPlayerGP,
            upsert_weekly_player_gp,
            mark_week_closed,
        )

        league_key = getattr(args, "league_key")
        start_week = int(getattr(args, "start_week", 1) or 1)
        end_week = int(getattr(args, "end_week", start_week) or start_week)
        team_keys_arg = getattr(args, "team_keys", None)
        batch_size = max(1, int(getattr(args, "batch_size", 3) or 3))
        sleep_sec = float(getattr(args, "sleep_sec", 2.0) or 0)
        force = bool(getattr(args, "force", False))
        no_close = bool(getattr(args, "no_close", False))
        dry_run = bool(getattr(args, "dry_run", False))

        if end_week < start_week:
            raise SystemExit("end-week must be >= start-week")

        session = get_session()
        try:
            league = session.query(League).filter(League.league_key == str(league_key)).one_or_none()
            if league is None:
                raise SystemExit(f"League not found: {league_key}")

            # Resolve team scope
            team_q = session.query(Team).filter(Team.league_id == league.id)
            if team_keys_arg:
                scope = [t.strip() for t in str(team_keys_arg).split(',') if t.strip()]
                team_q = team_q.filter(Team.team_key.in_(scope))
            teams = list(team_q.order_by(Team.team_key.asc()).all())
            if not teams:
                raise SystemExit("No teams found to process")

            total_written = 0
            for wk in range(start_week, end_week + 1):
                wk_row = session.query(Week).filter(Week.league_id == league.id, Week.week_num == int(wk)).one_or_none()
                if wk_row is None or not wk_row.start_date or not wk_row.end_date:
                    print(f"[Warn] Missing dates for week {wk}; skipping")
                    continue

                print(f"Processing Week {wk} ({wk_row.start_date}..{wk_row.end_date}) for {len(teams)} team(s)")

                # Preload existing WeeklyPlayerGP per team/week if skipping existing
                existing_by_team = {}
                if not force:
                    rows = (
                        session.query(WeeklyPlayerGP.team_key, WeeklyPlayerGP.player_key)
                        .filter(
                            WeeklyPlayerGP.league_id == league.id,
                            WeeklyPlayerGP.week_num == int(wk),
                        )
                        .all()
                    )
                    for tkey, pkey in rows:
                        existing_by_team.setdefault(tkey, set()).add(pkey)

                # Batch teams
                for i in range(0, len(teams), batch_size):
                    batch = teams[i:i + batch_size]
                    print(f"  Team batch {i//batch_size + 1}: {len(batch)} team(s)")
                    for t in batch:
                        tkey = t.team_key
                        # Aggregate daily gp by player
                        daily_q = (
                            session.query(RosterSlotDaily.player_key, func.coalesce(func.sum(RosterSlotDaily.gp), 0))
                            .filter(
                                RosterSlotDaily.league_id == league.id,
                                RosterSlotDaily.team_key == tkey,
                                RosterSlotDaily.date >= wk_row.start_date,
                                RosterSlotDaily.date <= wk_row.end_date,
                            )
                            .group_by(RosterSlotDaily.player_key)
                        )
                        agg = list(daily_q.all())
                        if not agg:
                            print(f"    {tkey}: no daily GP rows in date range; skipping")
                            continue

                        skip_set = existing_by_team.get(tkey, set()) if not force else set()
                        wrote = 0
                        for pkey, gp_sum in agg:
                            if (not force) and pkey in skip_set:
                                continue
                            if dry_run:
                                print(f"    [DRY] upsert_weekly_player_gp wk={wk} team={tkey} player={pkey} gp={int(gp_sum or 0)} source=daily_agg")
                                wrote += 1
                                continue
                            upsert_weekly_player_gp(
                                session,
                                league_id=league.id,
                                week_num=int(wk),
                                team_key=tkey,
                                player_key=str(pkey),
                                gp=int(gp_sum or 0),
                                source="daily_agg",
                                player_name=None,
                                positions=None,
                            )
                            wrote += 1
                        if wrote:
                            session.commit()
                            total_written += wrote
                            print(f"    {tkey}: wrote {wrote} WeeklyPlayerGP row(s)")
                        else:
                            print(f"    {tkey}: nothing to write (existing or empty)")

                    if sleep_sec > 0 and (i + batch_size) < len(teams):
                        time.sleep(sleep_sec)

                if not no_close and not dry_run:
                    try:
                        mark_week_closed(session, league_id=league.id, week_num=int(wk))
                        session.commit()
                        print(f"  Week {wk} marked closed")
                    except Exception as _e:
                        session.rollback()
                        print(f"  [Warn] Failed to mark week {wk} closed: {_e}")

            print(f"Done. Total WeeklyPlayerGP rows written: {total_written}")
        finally:
            try:
                session.close()
            except Exception:
                pass
    elif cmd == "fetch-daily-gp":
        # Build RosterSlotDaily from Yahoo per-date roster and stats (rate-limit aware)
        from yahoo.yahoo_auth import get_oauth
        from yahoo.yfs_client import YahooFantasyClient, parse_roster_xml, parse_settings_xml, parse_players_week_stats_xml
        from infrastructure.persistence import (
            get_session,
            League,
            Week,
            Team,
            upsert_team,
            upsert_roster_slot_daily,
        )

        league_key = getattr(args, "league_key")
        start_week = int(getattr(args, "start_week"))
        end_week = int(getattr(args, "end_week"))
        sleep_sec = float(getattr(args, "sleep_sec", 1.0) or 0)
        team_keys_arg = getattr(args, "team_keys", None)
        dry_run = bool(getattr(args, "dry_run", False))

        if end_week < start_week:
            raise SystemExit("end-week must be >= start-week")

        session = get_session()
        try:
            league = session.query(League).filter(League.league_key == str(league_key)).one_or_none()
            if league is None:
                raise SystemExit(f"League not found: {league_key}")

            # Resolve team scope from DB
            team_q = session.query(Team).filter(Team.league_id == league.id)
            if team_keys_arg:
                scope = [t.strip() for t in str(team_keys_arg).split(',') if t.strip()]
                team_q = team_q.filter(Team.team_key.in_(scope))
            teams = list(team_q.order_by(Team.team_key.asc()).all())
            if not teams:
                raise SystemExit("No teams found to process")

            # Resolve week date ranges
            weeks = (
                session.query(Week)
                .filter(Week.league_id == league.id, Week.week_num >= start_week, Week.week_num <= end_week)
                .order_by(Week.week_num.asc())
                .all()
            )
            week_map = {w.week_num: w for w in weeks if w.start_date and w.end_date}
            missing = [w for w in range(start_week, end_week + 1) if w not in week_map]
            if missing:
                print(f"[Warn] Missing dates for weeks: {missing}; they will be skipped")

            # Initialize Yahoo client with credentials from environment/.env
            # Load .env on demand to avoid requiring it for other commands
            try:
                from dotenv import load_dotenv  # type: ignore
                load_dotenv()
            except Exception:
                pass

            # Accept common env var spellings (case-insensitive) and guard for typos
            def _get_env_any(*keys):
                for k in keys:
                    v = os.getenv(k)
                    if v:
                        return v
                return None

            client_id = _get_env_any(
                "YAHOO_CLIENT_ID", "yahoo_client_id", "Yahoo_Client_Id"
            )
            client_secret = _get_env_any(
                "YAHOO_CLIENT_SECRET", "yahoo_client_secret", "Yahoo_Client_Secret",
                # also handle common typo without second 'i'
                "YAHOO_CLENT_SECRET", "yahoo_clent_secret"
            )
            if not client_id or not client_secret:
                raise SystemExit(
                    "Missing Yahoo OAuth credentials. Set YAHOO_CLIENT_ID and YAHOO_CLIENT_SECRET in your .env"
                )

            client = YahooFantasyClient(get_oauth(client_id, client_secret))

            # Discover GP stat id
            settings_payload = client.get_league_settings(league_key)
            settings_xml = settings_payload.get("_raw_xml") if isinstance(settings_payload, dict) else None
            gp_stat_id = None
            if settings_xml:
                settings = parse_settings_xml(settings_xml)
                for s in settings.get("stat_categories", []) or []:
                    try:
                        sid = int(s.get("id"))
                        disp = s.get("display") or s.get("name") or str(sid)
                        d = str(disp).strip().lower()
                        if gp_stat_id is None and (d == 'gp' or 'games played' in d):
                            gp_stat_id = sid
                    except Exception:
                        continue

            def daterange(start_d, end_d):
                cur = start_d
                while cur <= end_d:
                    yield cur
                    cur = cur + __import__('datetime').timedelta(days=1)

            total_rows = 0
            for wk in range(start_week, end_week + 1):
                wk_row = week_map.get(wk)
                if not wk_row:
                    print(f"[Skip] Week {wk} has no dates configured")
                    continue
                print(f"Processing Week {wk} ({wk_row.start_date}..{wk_row.end_date}) for {len(teams)} team(s)")

                for d in daterange(wk_row.start_date, wk_row.end_date):
                    date_iso = d.isoformat()
                    print(f"  Date {date_iso}")
                    for t in teams:
                        tkey = t.team_key
                        try:
                            roster_payload = client.get_team_roster_date(tkey, date_iso)
                            r_xml = roster_payload.get("_raw_xml") if isinstance(roster_payload, dict) else None
                            if not r_xml:
                                print(f"    [Warn] No roster XML for {tkey} on {date_iso}; skipping")
                                continue
                            roster = parse_roster_xml(r_xml)
                            team_name = roster.get('team_name') or roster.get('name') or ''
                            players = roster.get('players', []) or []
                            # Build player keys and meta
                            pkeys = []
                            pmeta = {}
                            for p in players:
                                pid = str(p.get('player_id') or '').strip()
                                if not pid.isdigit():
                                    continue
                                # player_key is game_key invariant; however Yahoo accepts any key for per-date stats
                                # We will build player_key based on the league's game part from team_key
                                # team_key looks like nhl.l.<league>.t.<id> → derive game_key = nhl
                                game_key = tkey.split('.l.')[0]
                                player_key = f"{game_key}.p.{pid}"
                                pkeys.append(player_key)
                                positions = p.get('positions') or []
                                pos_str = ",".join([str(x) for x in positions]) if isinstance(positions, (list, tuple)) else str(positions or '')
                                pname = p.get('name') or p.get('name_full') or ''
                                sel = (p.get('selected_position') or '').strip().upper()
                                pmeta[player_key] = {"name": pname, "positions": pos_str, "selected_position": sel}
                            if not pkeys:
                                continue

                            # Fetch player date stats in batches of 25
                            stats_by_key = {}
                            for i in range(0, len(pkeys), 25):
                                batch = pkeys[i:i+25]
                                try:
                                    pd_payload = client.get_players_date_stats(batch, date_iso)
                                    pd_xml = pd_payload.get("_raw_xml") if isinstance(pd_payload, dict) else None
                                    if not pd_xml:
                                        continue
                                    parsed = parse_players_week_stats_xml(pd_xml)
                                    for item in parsed.get('players', []) or []:
                                        stats_by_key[item.get('player_key')] = item.get('stats', {}) or {}
                                except Exception as _pe:
                                    print(f"    [Warn] player date stats failed {tkey} {date_iso}: {_pe}")
                                    continue

                            inactive_slots = {"BN", "IR", "IR+"}
                            # Upsert rows
                            wrote = 0
                            for pkey in pkeys:
                                meta = pmeta.get(pkey, {})
                                stats = stats_by_key.get(pkey, {}) or {}
                                # had_game detection using GP stat if available, else any positive stat
                                gp_val = None
                                if gp_stat_id is not None and gp_stat_id in stats:
                                    gp_val = stats.get(gp_stat_id)
                                elif 0 in stats:
                                    gp_val = stats.get(0)
                                had_game = False
                                if gp_val is not None:
                                    try:
                                        had_game = int(float(str(gp_val))) > 0
                                    except Exception:
                                        had_game = False
                                else:
                                    for v in stats.values():
                                        try:
                                            if float(str(v)) > 0:
                                                had_game = True
                                                break
                                        except Exception:
                                            continue
                                sel_pos = (meta.get('selected_position') or '').strip().upper()
                                active_slot = (sel_pos != '' and sel_pos not in inactive_slots)
                                gp_num = 1 if (active_slot and had_game) else 0

                                if dry_run:
                                    print(f"    [DRY] upsert RSD {date_iso} team={tkey} player={pkey} sel={sel_pos} had_game={had_game} gp={gp_num}")
                                    wrote += 1
                                    continue

                                # Ensure team row and then upsert daily
                                upsert_team(session, league_id=league.id, team_key=tkey, team_name=team_name)
                                upsert_roster_slot_daily(
                                    session,
                                    date=d,
                                    league_id=league.id,
                                    team_key=tkey,
                                    player_key=pkey,
                                    selected_position=sel_pos,
                                    had_game=bool(had_game),
                                    gp=int(gp_num),
                                    player_name=meta.get('name') or None,
                                    positions=meta.get('positions') or None,
                                )
                                wrote += 1
                            if wrote and not dry_run:
                                session.commit()
                                total_rows += wrote
                                # Gentle sleep between teams to avoid rate limits
                            if sleep_sec > 0:
                                time.sleep(sleep_sec)
                        except Exception as _te:
                            print(f"    [Warn] Failed team {tkey} on {date_iso}: {_te}")
                            continue

            print(f"Done. RosterSlotDaily rows written: {total_rows}")
        finally:
            try:
                session.close()
            except Exception:
                pass
    elif cmd == "init-weeks":
        # Populate Week table (start/end dates) from nhl_schedule (Yahoo-defined week windows)
        import csv
        import datetime as _dt
        import os as _os
        from infrastructure.persistence import get_session, League, upsert_week

        league_key = getattr(args, "league_key")
        start_week = int(getattr(args, "start_week", 10) or 10)
        end_week = int(getattr(args, "end_week", 24) or 24)
        refresh_cache = bool(getattr(args, "refresh_cache", False))
        source = str(getattr(args, "source", "auto") or "auto").lower()
        excel_path = getattr(args, "excel_path", None)
        excel_sheet = getattr(args, "excel_sheet", "Yahoo_Weeks")
        debug = bool(getattr(args, "debug", False))
        dry_run = bool(getattr(args, "dry_run", False))

        if end_week < start_week:
            raise SystemExit("end-week must be >= start-week")

        # Resolve season: arg --season overrides .env SEASON
        season = getattr(args, "season", None)
        if not season:
            try:
                from dotenv import load_dotenv  # type: ignore
                load_dotenv()
            except Exception:
                pass
            season = _os.getenv("SEASON")
        if not season:
            raise SystemExit("Season not provided. Pass --season or set SEASON in .env")
        season = str(season).strip()

        # Attempt to build a season lookup CSV via nhl_schedule (for lookup/auto)
        out_csv_path = os.path.join("data", f"lookup_table_{season}.csv")
        csv_ready = False
        if source in ("auto", "lookup"):
            try:
                from nhl_schedule.build_lookup import build as _build_lookup
            except Exception as e:
                if source == "lookup":
                    raise SystemExit(
                        "Failed to import nhl_schedule. Install the package or set NHL_SCHEDULE_PATH, or run schedule-lookup first."
                    ) from e
                if debug:
                    print("[init-weeks] nhl_schedule not importable; skipping CSV source")
            else:
                try:
                    _ = _build_lookup(
                        schedule_path=None,
                        sheet_or_table=None,
                        out_csv=out_csv_path,
                        out_xlsx=None,
                        refresh_cache=refresh_cache,
                        season=str(season) if "season" in _build_lookup.__code__.co_varnames else None,
                    )
                except TypeError:
                    _ = _build_lookup(
                        schedule_path=None,
                        sheet_or_table=None,
                        out_csv=out_csv_path,
                        out_xlsx=None,
                        refresh_cache=refresh_cache,
                    )
                csv_ready = True

        # Parse the lookup CSV for Yahoo week numbers and dates
        def _parse_date(s: str) -> _dt.date:
            s = (s or "").strip()
            try:
                return _dt.date.fromisoformat(s)
            except Exception:
                for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(s, fmt).date()
                    except Exception:
                        pass
            raise SystemExit(f"Invalid date value in schedule: {s}")

        with open(out_csv_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise SystemExit("Schedule CSV appears empty; cannot derive weeks")
            # Normalize headers, build lookup of normalized->original
            raw_headers = list(reader.fieldnames)
            norm_headers = [(h or "").strip().lstrip("\ufeff") for h in raw_headers]
            keys = [h.lower().replace(" ", "_") for h in norm_headers]
            header_map = {k: h for k, h in zip(keys, raw_headers)}

            # Aliases
            def pick(*cands):
                for c in cands:
                    if c in header_map:
                        return header_map[c]
                return None

            week_col = pick("week", "week_num", "weeknum", "week_id", "weekid", "yahoo_week")
            date_col = pick("date", "game_date", "dt")
            start_col = pick("start", "start_date", "startdate")
            end_col = pick("end", "end_date", "enddate")

            if debug:
                print(f"[init-weeks] Detected headers: {raw_headers}")
                print(f"[init-weeks] Mapped columns -> week:{week_col} date:{date_col} start:{start_col} end:{end_col}")

            wk_bounds: dict[int, tuple[_dt.date, _dt.date]] = {}

            if week_col and start_col and end_col:
                # Weekly summary shape (prefer this if available)
                for row in reader:
                    try:
                        wk = int(str(row.get(week_col, "")).strip())
                    except Exception:
                        continue
                    sd_raw = str(row.get(start_col, "")).strip()
                    ed_raw = str(row.get(end_col, "")).strip()
                    if not sd_raw or not ed_raw:
                        continue
                    sd = _parse_date(sd_raw)
                    ed = _parse_date(ed_raw)
                    if wk not in wk_bounds:
                        wk_bounds[wk] = (sd, ed)
                    else:
                        # In case multiple rows, keep min/max
                        lo, hi = wk_bounds[wk]
                        wk_bounds[wk] = (sd if sd < lo else lo, ed if ed > hi else hi)
            elif week_col and date_col:
                # Per-date shape: compute min/max per week
                for row in reader:
                    try:
                        wk = int(str(row.get(week_col, "")).strip())
                    except Exception:
                        continue
                    try:
                        d = _parse_date(str(row.get(date_col, "")).strip())
                    except SystemExit:
                        continue
                    if wk in wk_bounds:
                        lo, hi = wk_bounds[wk]
                        if d < lo:
                            lo = d
                        if d > hi:
                            hi = d
                        wk_bounds[wk] = (lo, hi)
                    else:
                        wk_bounds[wk] = (d, d)
            else:
                # CSV headers missing expected shapes -> optionally fall back to Excel
                if source == "lookup":
                    detected = ", ".join(raw_headers)
                    raise SystemExit(
                        "Schedule CSV is missing required columns. Expected either: "
                        "(week + date) or (week + start + end). Detected headers: " + detected
                    )
                # Try Excel if allowed
                wk_bounds = {}
                tried_excel = False
                if source in ("auto", "excel"):
                    tried_excel = True
                    x_path = excel_path or _os.getenv("NHL_SCHEDULE_XLSX_PATH")
                    if not x_path or not _os.path.exists(x_path):
                        if debug:
                            print(f"[init-weeks] Excel path not provided or not found: {x_path}")
                    else:
                        try:
                            from openpyxl import load_workbook  # type: ignore
                        except Exception as e:
                            raise SystemExit("openpyxl is required to read Excel. Install with: pip install openpyxl") from e
                        wb = load_workbook(x_path, data_only=True, read_only=True)
                        if excel_sheet not in wb.sheetnames:
                            raise SystemExit(f"Excel sheet not found: {excel_sheet}. Available: {wb.sheetnames}")
                        ws = wb[excel_sheet]
                        rows_iter = ws.iter_rows(values_only=True)
                        try:
                            header_row = next(rows_iter)
                        except StopIteration:
                            raise SystemExit("Excel sheet is empty; cannot derive weeks")
                        raw_xhdr = [str(h or '').strip() for h in header_row]
                        norm = [h.strip().lstrip("\ufeff").lower().replace(' ', '_') for h in raw_xhdr]
                        idx_map = {k: i for i, k in enumerate(norm)}
                        def _idx(*cands):
                            for c in cands:
                                k = c.strip().lower().replace(' ', '_')
                                if k in idx_map:
                                    return idx_map[k]
                            return None
                        i_week = _idx("week", "week_num", "weekid", "week_id")
                        i_start = _idx("start", "start_date", "startdate")
                        i_end = _idx("end", "end_date", "enddate")
                        if debug:
                            print(f"[init-weeks] Excel headers: {raw_xhdr}")
                            print(f"[init-weeks] Excel mapped idx -> week:{i_week} start:{i_start} end:{i_end}")
                        if i_week is None or i_start is None or i_end is None:
                            raise SystemExit("Excel sheet must include columns for Week/Start/End")
                        # Parse rows
                        for r in rows_iter:
                            try:
                                wv = r[i_week]
                                if wv is None or str(wv).strip() == '':
                                    continue
                                wk = int(str(wv).strip())
                            except Exception:
                                continue
                            s_raw = r[i_start]
                            e_raw = r[i_end]
                            try:
                                sd = _parse_date(s_raw.isoformat() if hasattr(s_raw, 'isoformat') else str(s_raw))
                                ed = _parse_date(e_raw.isoformat() if hasattr(e_raw, 'isoformat') else str(e_raw))
                            except SystemExit:
                                continue
                            wk_bounds[wk] = (sd, ed)
                        if not wk_bounds:
                            raise SystemExit("No week rows parsed from Excel source")
                        # Continue after Excel fallback
                if not wk_bounds:
                    detected = ", ".join(raw_headers)
                    tried = [f"CSV:{out_csv_path}"]
                    if tried_excel:
                        tried.append(f"Excel:{excel_path or _os.getenv('NHL_SCHEDULE_XLSX_PATH')}")
                    raise SystemExit(
                        "Unable to derive week windows from provided sources. Tried " + ", ".join(tried) + ". "
                        "CSV detected headers: " + detected + ". Use --excel-path/--excel-sheet or --debug for details."
                    )

        # Select desired week range
        selected = [(wk, lo, hi) for wk, (lo, hi) in wk_bounds.items() if start_week <= wk <= end_week]
        if not selected:
            raise SystemExit(f"No weeks found in range {start_week}..{end_week} for season {season}")
        selected.sort(key=lambda x: x[0])

        if debug:
            preview = ", ".join([f"w{wk}:{lo}..{hi}" for wk, lo, hi in selected[:3]])
            print(f"[init-weeks] Parsed weeks preview: {preview}")

        session = get_session()
        try:
            league = session.query(League).filter(League.league_key == str(league_key)).one_or_none()
            if league is None:
                raise SystemExit(f"League not found: {league_key}")

            wrote = 0
            for wk, sd, ed in selected:
                if dry_run:
                    print(f"[DRY] upsert_week league={league.league_key} week={wk} {sd}..{ed}")
                    wrote += 1
                    continue
                upsert_week(session, league_id=league.id, week_num=int(wk), start_date=sd, end_date=ed)
                wrote += 1
            if not dry_run:
                session.commit()
            print(f"init-weeks: processed {len(selected)} week(s); {'planned' if dry_run else 'upserted'} {wrote}")
        finally:
            try:
                session.close()
            except Exception:
                pass
    elif cmd == "waiver-agent":
        team_name = getattr(args, "team_name")
        current_week = int(getattr(args, "current_week"))
        skaters_csv = getattr(args, "skaters_csv", os.path.join("data","merged_skaters.csv"))
        lookup_csv = getattr(args, "lookup_csv", os.path.join("data","lookup_table.csv"))
        prior_csv = getattr(args, "prior_csv", os.path.join("data","merged_skaters_prior.csv"))
        out_dir = getattr(args, "out_dir", os.path.join("output"))
        paths = run_waiver(
            team_name=team_name,
            current_week=current_week,
            skaters_csv=skaters_csv,
            lookup_csv=lookup_csv,
            prior_csv=prior_csv,
            out_dir=out_dir,
        )
        print("Waiver recommendations written to:")
        print(f"  CSV: {paths.get('csv')}")
        print(f"  Plots: {paths.get('plots_dir')}")
    else:
        run_all()


if __name__ == "__main__":
    main()
